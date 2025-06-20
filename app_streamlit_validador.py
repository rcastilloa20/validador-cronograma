
import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Data Entry Cronograma de inventario", layout="centered")

st.title("📦 Data Entry - Cronograma de inventario")
st.markdown("Valida tipos de datos en tu cronograma de inventario y genera reportes automáticos.")

uploaded_file = st.file_uploader("Selecciona tu archivo CSV", type=["csv"])

def es_fecha_valida(valor):
    return bool(re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', str(valor).strip()))

def es_entero_valido(valor):
    val = str(valor).strip()
    return val.isdigit() or (val.startswith('-') and val[1:].isdigit())

def es_float_valido(valor):
    try:
        float(str(valor).replace(',', '.'))
        return True
    except:
        return False

def procesar_dataframe(df):
    columnas_fecha_ddmm = [
        'fecha_apertura', 'fecha_cierre', 'exhib_nutricion', 'f_implem', 'f_baja_implem',
        'ultim_inv_2022', 'f_inv_2022', 'f_maxima', 'f_inv_gestion', 'f_inv_e_sist',
        'f_inv_cierre', 'f_inv_siniestro', 'f_inv_gral', 'inv_gond_enero', 'inv_gond_febrero',
        'inv_gond_marzo', 'inv_gond_abril', 'inv_gond_mayo', 'inv_gond_junio', 'inv_gond_julio',
        'inv_gond_agosto', 'inv_gond_setiembre', 'inv_gond_octubre', 'inv_gond_noviembre',
        'inv_gond_diciembre'
    ]
    columnas_float = ['stock']
    columnas_int = ['ceco_soc_a111', 'rpc', 'q_inv', 'total_inv']

    resultado = []
    datos_inconsistentes = {}

    for col in df.columns:
        total_valores = len(df[col])
        nulos = df[col].isna().sum()
        valores = df[col].dropna()

        if col in columnas_fecha_ddmm:
            tipo_esperado = 'DATE'
            validador = es_fecha_valida
        elif col in columnas_int:
            tipo_esperado = 'INT64'
            validador = es_entero_valido
        elif col in columnas_float:
            tipo_esperado = 'FLOAT64'
            validador = es_float_valido
        else:
            tipo_esperado = 'STRING'
            validador = lambda x: True

        inconsistencias_idx = [i for i, val in valores.items() if not validador(val)]
        inconsistencias = len(inconsistencias_idx)

        if inconsistencias > 0:
            datos_inconsistentes[col] = df.loc[inconsistencias_idx, [col]]

        resultado.append({
            'Columna': col,
            'Tipo esperado': tipo_esperado,
            'Total valores': total_valores,
            'Valores vacíos (NaN)': nulos,
            'Total valores no nulos': total_valores - nulos,
            'Valores inconsistentes': inconsistencias,
            'Porcentaje inconsistencias': round((inconsistencias / (total_valores - nulos)) * 100, 4) if (total_valores - nulos) > 0 else 0.0
        })

    return resultado, datos_inconsistentes

def generar_reporte_excel(resumen, inconsistencias):
    output = BytesIO()
    df = pd.DataFrame(resumen)
    df.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    fill_cabecera = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    font_blanca = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill_cabecera
        cell.font = font_blanca

    for row in range(2, ws.max_row + 1):
        ws[f'G{row}'].number_format = '0.00%'
        inconsistencias_val = ws[f'F{row}'].value
        if inconsistencias_val > 0:
            ws[f'F{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws[f'G{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = False

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

if uploaded_file:
    try:
        df = pd.read_csv(uploaded_file, encoding="latin1", sep=";", dtype=str, on_bad_lines='skip')
        st.success("✅ Archivo cargado correctamente.")
        resumen, inconsistencias = procesar_dataframe(df)

        total_errores = sum([r["Valores inconsistentes"] for r in resumen])
        if total_errores > 0:
            st.markdown(f"""
                <div style="
                    position: fixed;
                    top: 30%;
                    left: 50%;
                    transform: translate(-50%, -50%);
                    background-color: #FFF3CD;
                    padding: 30px;
                    border-left: 6px solid #FFA500;
                    border-radius: 12px;
                    font-size: 20px;
                    font-weight: bold;
                    color: #856404;
                    box-shadow: 0px 4px 20px rgba(0,0,0,0.3);
                    z-index: 9999;
                    width: 80%;
                    max-width: 600px;
                    text-align: center;
                ">
                    ⚠️ Se encontraron {total_errores} valores inconsistentes en el archivo cargado.
                </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
                <div style="
                    position: fixed;
                    top: 30%;
                    left: 50%;
                    transform: translate(-50%, -50%);
                    background-color: #D4EDDA;
                    padding: 30px;
                    border-left: 6px solid #28A745;
                    border-radius: 12px;
                    font-size: 20px;
                    font-weight: bold;
                    color: #155724;
                    box-shadow: 0px 4px 20px rgba(0,0,0,0.3);
                    z-index: 9999;
                    width: 80%;
                    max-width: 600px;
                    text-align: center;
                ">
                    ✅ No se encontraron inconsistencias en los datos.
                </div>
            """, unsafe_allow_html=True)

        st.subheader("🔍 Resumen de validación")

        resumen_df = pd.DataFrame(resumen)

        def resaltar_errores(row):
            color = 'background-color: #FFC7CE' if row['Valores inconsistentes'] > 0 else ''
            return [color] * len(row)

        styled_df = resumen_df.style.apply(resaltar_errores, axis=1)
        st.dataframe(styled_df, use_container_width=True)

        excel_report = generar_reporte_excel(resumen, inconsistencias)

        st.download_button(
            label="📥 Descargar reporte Excel",
            data=excel_report,
            file_name="reporte_validacion.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if inconsistencias:
            inconsistencias_df = pd.concat(inconsistencias.values(), axis=1)
            inconsistencias_xlsx = BytesIO()
            inconsistencias_df.to_excel(inconsistencias_xlsx, index=False)
            inconsistencias_xlsx.seek(0)

            st.download_button(
                label="❗ Descargar solo datos inconsistentes",
                data=inconsistencias_xlsx,
                file_name="datos_inconsistentes.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"❌ Error al procesar el archivo: {e}")
